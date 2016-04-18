#!/usr/bin/python

import os
import sys
import json
import shutil
import zipfile
import xlrd
from xlutils.copy import copy
from copy import deepcopy
from openpyxl.workbook import Workbook
import pdb


"""
INPUTS: 1) Reference XLSM template (does this have 1:1 sheets, or blank?)
        2) config.json that needs to be converted
OUTPUTS: 1) config.xlsm
STEPS:  1) Unpack the xlsm file that we're using as our reference/template into a temp directory.
        2) Open the config.json that we want to convert over to xlsm.
        3) Create an xlsx file first by looking at every value in the reference and replacing with the value in the json.
           Note that we don't attempt to add any value from the json that isn't in the reference because we'd have no idea
           what sheet to put it in.
        4) Save the xlsx file.
        5) Unpack the xlsx file into a (second) temp directory.
        6) Copy all the worksheet.xml files from the unpacked xlsx into the xlsm unpacked directory (EXCEPT worksheet7.xml!)
        7) Zip up the xlsm unpacked directory into a new xlsm. This is our final product.
"""

if len( sys.argv ) < 3:
    print( "Usage: json_to_xlsm.py <reference.xlsm> <sim_dir>" )
    sys.exit()

config_json_path = os.path.join( sys.argv[2], "config.json" )
if os.path.exists( config_json_path ) == False:
    print( "Couldn't find 'config.json' in specified directory: " + sys.argv[2] );
    sys.exit()

config_xlsm_path = sys.argv[1]
if os.path.exists( config_xlsm_path ) == False:
    print( "Couldn't find {0} as specified." ).format( config_xlsm_path )
    sys.exit()

# STEP 1) Unpack the xlsm file that we're using as our reference/template into a temp directory.
print( "Unpacking xlsm file..." )
outpath = ".xlsm_unpack_tmp"
if os.path.exists( outpath ):
    shutil.rmtree( outpath )
else:
    os.mkdir( outpath  )

with zipfile.ZipFile(config_xlsm_path, "r") as z:
    z.extractall( outpath )

# STEP 2) Open the config.json that we want to convert over to xlsm.
cj = json.loads( open( config_json_path ).read() )["parameters"]
print( "Converting json to xlsx file...." )

control_sheet = 0
sheet_num = 0
# STEP 3) Create an xlsx file first by looking at every value in the reference workbook and replacing with the value in the json.
wb = xlrd.open_workbook( sys.argv[1] ) 
for sheet in wb.sheets():
    sheet_num = sheet_num + 1
    if sheet.name=="CONTROL":
        control_sheet = sheet_num
    else:
        if sheet.name.startswith( "STI_Network_Params" ):
            #pdb.set_trace()
            # need special handling
            sti_param_group = sheet.name.split( "-" )[2] 
            if sti_param_group not in cj["STI_Network_Params_By_Property"]:
                print( "Yikes! config.json doesn't have section to override the STI_Network_Params property. Nothing will be changed here." )
            else:
                use_node = cj["STI_Network_Params_By_Property"][ sti_param_group ]
                for row_id in range(0,sheet.nrows):
                    row = sheet.row(row_id)
                    param_name = row[0].value
                    if param_name in use_node:
                        if row[1].value != use_node[ param_name ]:
                            if not isinstance(use_node[param_name],list):
                                #print( "Replacing default value for " + param_name + " with config.json value of " + str( use_node[ param_name ] ) )
                                sheet._cell_values[row_id][1] = use_node[ param_name ]
                                print( "Replaced default value for " + param_name + " with config.json value of " + str( sheet.row(row_id)[1].value ) )
                            #else:
                                #sheet._cell_values[row_id][1] = str(use_node[ param_name ])

        for row_id in range(0,sheet.nrows):
            row = sheet.row(row_id)
            param_name = row[0].value
            # param_value = row[1].value
            # find param_name in config.json
            if param_name in cj:
                if row[1].value != cj[ param_name ]:
                    if not isinstance(cj[param_name],list):
                        #print( "Replacing default value for " + param_name + " with config.json value of " + str( cj[ param_name ] ) )
                        sheet._cell_values[row_id][1] = cj[ param_name ]
                        print( "Replaced default value for " + param_name + " with config.json value of " + str( sheet.row(row_id)[1].value ) )
                    else:
                        sheet._cell_values[row_id][1] = str(cj[ param_name ])
            elif param_name == "Sim Path":
                sheet._cell_values[row_id][1] = sys.argv[2].replace( "Scenarios\\", "" )
                print( "Using Sim Path: " + sheet._cell_values[row_id][1] )

xlsBook = wb
workbook = Workbook()

for i in xrange(0, xlsBook.nsheets):
    xlsSheet = xlsBook.sheet_by_index(i)
    sheet = workbook.active if i == 0 else workbook.create_sheet()
    sheet.title = xlsSheet.name

    for r in xrange(0, xlsSheet.nrows):
        for col in xrange(0, xlsSheet.ncols):
            sheet.cell(row=r+1, column=col+1).value = xlsSheet.cell_value(r, col)

# STEP 4) Save the xlsx file.
xlsx_path = os.path.join( sys.argv[2], "config_new.xlsx" )
workbook.save( os.path.join( xlsx_path ) )

# STEP 5) Unpack the xlsx file into a (second) temp directory.
print( "Unpacking xlsx file..." )
xlsx_outpath = ".xlsx_unpack_tmp"
if os.path.exists( xlsx_outpath ):
    shutil.rmtree( xlsx_outpath )
else:
    os.mkdir( xlsx_outpath  )

with zipfile.ZipFile(xlsx_path, "r") as z:
    z.extractall( xlsx_outpath )

# STEP 6) Copy all the worksheet.xml files from the unpacked xlsx into the xlsm unpacked directory (EXCEPT worksheet7.xml!)
subdir_path = os.path.join( xlsx_outpath, "xl/worksheets" )
for ws_file in os.listdir( subdir_path ):
    if ws_file.startswith( "sheet" ) and ws_file.endswith( ".xml" ):
        if ws_file == "sheet" + str(control_sheet) +".xml":
            continue
        src_path = os.path.join( subdir_path, ws_file )
        dest_path = os.path.join( outpath, "xl/worksheets", ws_file )
        print( "Copying {0} to {1}." ).format( src_path, dest_path )
        shutil.copy( src_path, dest_path )
# copy strings file
shutil.copy( os.path.join( xlsx_outpath, "xl/sharedStrings.xml" ), os.path.join( outpath, "xl/sharedStrings.xml" ) )

# STEP 6b) Replace config name with new name in sheet7.xml (the CONTROL sheet).

strings_filepath = os.path.join( outpath, "xl/sharedStrings.xml" )
strings_file = open( strings_filepath )
data = strings_file.read()
strings_file.close()
data = data.replace( "CONFIG_PATH", sys.argv[2] )
strings_file = open( strings_filepath, "w" )
strings_file.write( data )
strings_file.close()

# STEP 7) Zip up the xlsm unpacked directory into a new xlsm. This is our final product.

output_filename = "config.xlsm"
outputfile = os.path.join( os.getcwd(), os.path.join( sys.argv[2], output_filename ) )
print( "Creating " + outputfile )
os.chdir( outpath )
shutil.make_archive( outputfile, 'zip', "." )
os.remove( outputfile )
os.rename( outputfile + ".zip", outputfile )
os.chdir( ".." )
shutil.rmtree( outpath )
shutil.rmtree( xlsx_outpath )
